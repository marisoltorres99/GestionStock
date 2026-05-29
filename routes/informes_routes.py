from flask import Blueprint, render_template
from models.venta import Venta
from models.detalle_venta import DetalleVenta
from models.producto import Producto
from models.cliente import Cliente
from datetime import datetime, timedelta
from sqlalchemy import func, select
from models.db import db
from services.deuda_service import calcular_deuda_cliente

from openpyxl import Workbook
from flask import send_file
from io import BytesIO

from reportlab.platypus import (
    SimpleDocTemplate,
    Table,
    TableStyle,
    Paragraph,
    Spacer
)

from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.pagesizes import letter

informe_bp = Blueprint("informe", __name__)

@informe_bp.route("/informes")
def listar():
    return render_template("informes/index.html")

@informe_bp.route("/informes/ventasDelDia")
def ventasDelDia():

    inicio_hoy = datetime.today().replace(
    hour=0,
    minute=0,
    second=0,
    microsecond=0
    )

    inicio_manana = inicio_hoy + timedelta(days=1)
    
    ventasDiarias = Venta.query.filter(Venta.fecha >= inicio_hoy, Venta.fecha < inicio_manana).all()

    totalRecaudadoHoy = (
        db.session.query(func.sum(Venta.total))
        .filter(
            Venta.fecha >= inicio_hoy,
            Venta.fecha < inicio_manana
        )
        .scalar()
    ) or 0

    return render_template("informes/ventasDiarias.html", totalRecaudadoHoy=totalRecaudadoHoy, ventas=ventasDiarias)

@informe_bp.route("/informes/productosMasVendidos")
def productosMasVendidos():
    
    stmt = (
        select(Producto.nombre, func.sum(DetalleVenta.cantidad).label("cantidadProductoVendido"))
        .join(DetalleVenta, Producto.id == DetalleVenta.producto_id)
        .group_by(Producto.id, Producto.nombre)
        .order_by(func.sum(DetalleVenta.cantidad).desc())
    )

    detallesVentas = db.session.execute(stmt).all()

    return render_template("informes/productosMasVendidos.html", detallesVentas=detallesVentas)

@informe_bp.route("/informes/clientesConDeudas")
def clientesConDeudas():
    
    clientes = db.session.query(Cliente).all()

    clientes_con_deudas = []

    for cliente in clientes:
        monto_ventas_fiadas, monto_pagos, deuda = calcular_deuda_cliente(cliente.id)
        if deuda > 0:
            clientes_con_deudas.append({
                "cliente": cliente,
                "deuda": deuda,
                "pagos": monto_pagos,
                "ventas": monto_ventas_fiadas
            })
    
    clientes_con_deudas.sort(
        key=lambda x: x["deuda"],
        reverse=True
    )

    return render_template("informes/clientesConDeudas.html", clientes_con_deudas=clientes_con_deudas)

@informe_bp.route("/informes/ventasDelDia/excel")
def exportarVentasExcel():

    inicio_hoy = datetime.today().replace(
        hour=0,
        minute=0,
        second=0,
        microsecond=0
    )

    inicio_manana = inicio_hoy + timedelta(days=1)

    ventas = Venta.query.filter(
        Venta.fecha >= inicio_hoy,
        Venta.fecha < inicio_manana
    ).all()

    wb = Workbook()

    ws = wb.active

    ws.title = "Ventas del Día"

    ws.append([
        "ID",
        "Fecha",
        "Cliente",
        "Tipo Pago",
        "Total"
    ])

    for venta in ventas:

        cliente = (
            venta.cliente.nombre
            if venta.cliente
            else "Consumidor Final"
        )

        ws.append([
            venta.id,
            venta.fecha.strftime('%d/%m/%Y %H:%M'),
            cliente,
            venta.tipo_pago,
            venta.total
        ])

    archivo = BytesIO()

    wb.save(archivo)

    archivo.seek(0)

    return send_file(
        archivo,
        as_attachment=True,
        download_name="ventas_del_dia.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

@informe_bp.route("/informes/ventasDelDia/pdf")
def exportarVentasPDF():

    inicio_hoy = datetime.today().replace(
        hour=0,
        minute=0,
        second=0,
        microsecond=0
    )

    inicio_manana = inicio_hoy + timedelta(days=1)

    ventas = Venta.query.filter(
        Venta.fecha >= inicio_hoy,
        Venta.fecha < inicio_manana
    ).all()

    buffer = BytesIO()

    doc = SimpleDocTemplate(
        buffer,
        pagesize=letter
    )

    elementos = []

    estilos = getSampleStyleSheet()

    titulo = Paragraph(
        "Reporte Ventas del Día",
        estilos['Title']
    )

    elementos.append(titulo)

    elementos.append(Spacer(1, 20))

    datos = [[
        "ID",
        "Fecha",
        "Cliente",
        "Tipo Pago",
        "Total"
    ]]

    for venta in ventas:

        cliente = (
            venta.cliente.nombre
            if venta.cliente
            else "Consumidor Final"
        )

        datos.append([
            venta.id,
            venta.fecha.strftime('%d/%m/%Y %H:%M'),
            cliente,
            venta.tipo_pago,
            f"$ {venta.total}"
        ])

    tabla = Table(datos)

    tabla.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.grey),
        ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke),
        ('GRID', (0,0), (-1,-1), 1, colors.black),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0,0), (-1,0), 10),
    ]))

    elementos.append(tabla)

    doc.build(elementos)

    buffer.seek(0)

    return send_file(
        buffer,
        as_attachment=True,
        download_name="ventas_del_dia.pdf",
        mimetype="application/pdf"
    )

@informe_bp.route("/informes/clientesDeudas/excel")
def exportarClientesDeudasExcel():

    clientes = db.session.query(Cliente).all()

    clientes_con_deudas = []

    for cliente in clientes:

        monto_ventas_fiadas, monto_pagos, deuda = calcular_deuda_cliente(cliente.id)

        if deuda > 0:

            clientes_con_deudas.append({
                "cliente": cliente,
                "deuda": deuda,
                "pagos": monto_pagos,
                "ventas": monto_ventas_fiadas
            })

    wb = Workbook()

    ws = wb.active

    ws.title = "Clientes con Deudas"

    ws.append([
        "Cliente",
        "Ventas Fiadas",
        "Pagos",
        "Deuda"
    ])

    for item in clientes_con_deudas:

        ws.append([
            item["cliente"].nombre,
            item["ventas"],
            item["pagos"],
            item["deuda"]
        ])

    archivo = BytesIO()

    wb.save(archivo)

    archivo.seek(0)

    return send_file(
        archivo,
        as_attachment=True,
        download_name="clientes_con_deudas.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

@informe_bp.route("/informes/clientesDeudas/pdf")
def exportarClientesDeudasPDF():

    clientes = db.session.query(Cliente).all()

    clientes_con_deudas = []

    for cliente in clientes:

        monto_ventas_fiadas, monto_pagos, deuda = calcular_deuda_cliente(cliente.id)

        if deuda > 0:

            clientes_con_deudas.append({
                "cliente": cliente,
                "deuda": deuda,
                "pagos": monto_pagos,
                "ventas": monto_ventas_fiadas
            })

    buffer = BytesIO()

    doc = SimpleDocTemplate(
        buffer,
        pagesize=letter
    )

    elementos = []

    estilos = getSampleStyleSheet()

    titulo = Paragraph(
        "Reporte Clientes con Deudas",
        estilos['Title']
    )

    elementos.append(titulo)

    elementos.append(Spacer(1, 20))

    datos = [[
        "Cliente",
        "Ventas Fiadas",
        "Pagos",
        "Deudas"
    ]]

    for item in clientes_con_deudas:

        datos.append([
            item["cliente"].nombre,
            f"$ {item['ventas']}",
            f"$ {item['pagos']}",
            f"$ {item['deuda']}"
        ])

    tabla = Table(datos)

    tabla.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.grey),
        ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke),
        ('GRID', (0,0), (-1,-1), 1, colors.black),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0,0), (-1,0), 10),
    ]))

    elementos.append(tabla)

    doc.build(elementos)

    buffer.seek(0)

    return send_file(
        buffer,
        as_attachment=True,
        download_name="clientes_con_deudas.pdf",
        mimetype="application/pdf"
    )

@informe_bp.route("/informes/productosMasVendidos/excel")
def exportarProductosMasVendidosExcel():

    stmt = (
        select(
            Producto.nombre,
            func.sum(DetalleVenta.cantidad).label("cantidadProductoVendido")
        )
        .join(
            DetalleVenta,
            Producto.id == DetalleVenta.producto_id
        )
        .group_by(
            Producto.id,
            Producto.nombre
        )
        .order_by(
            func.sum(DetalleVenta.cantidad).desc()
        )
    )

    detallesVentas = db.session.execute(stmt).all()

    wb = Workbook()

    ws = wb.active

    ws.title = "Productos Más Vendidos"

    ws.append([
        "Producto",
        "Cantidad Vendida"
    ])

    for detalle in detallesVentas:

        ws.append([
            detalle.nombre,
            detalle.cantidadProductoVendido
        ])

    archivo = BytesIO()

    wb.save(archivo)

    archivo.seek(0)

    return send_file(
        archivo,
        as_attachment=True,
        download_name="productos_mas_vendidos.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

@informe_bp.route("/informes/productosMasVendidos/pdf")
def exportarProductosMasVendidosPDF():

    stmt = (
        select(
            Producto.nombre,
            func.sum(DetalleVenta.cantidad).label("cantidadProductoVendido")
        )
        .join(
            DetalleVenta,
            Producto.id == DetalleVenta.producto_id
        )
        .group_by(
            Producto.id,
            Producto.nombre
        )
        .order_by(
            func.sum(DetalleVenta.cantidad).desc()
        )
    )

    detallesVentas = db.session.execute(stmt).all()

    buffer = BytesIO()

    doc = SimpleDocTemplate(
        buffer,
        pagesize=letter
    )

    elementos = []

    estilos = getSampleStyleSheet()

    titulo = Paragraph(
        "Reporte Productos Más Vendidos",
        estilos['Title']
    )

    elementos.append(titulo)

    elementos.append(Spacer(1, 20))

    datos = [[
        "Producto",
        "Cantidad Vendida"
    ]]

    for detalle in detallesVentas:

        datos.append([
            detalle.nombre,
            detalle.cantidadProductoVendido
        ])

    tabla = Table(datos)

    tabla.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.grey),
        ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke),
        ('GRID', (0,0), (-1,-1), 1, colors.black),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0,0), (-1,0), 10),
    ]))

    elementos.append(tabla)

    doc.build(elementos)

    buffer.seek(0)

    return send_file(
        buffer,
        as_attachment=True,
        download_name="productos_mas_vendidos.pdf",
        mimetype="application/pdf"
    )